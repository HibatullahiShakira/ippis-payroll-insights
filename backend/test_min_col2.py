import openpyxl
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
wb.remove(ws)
ws = wb.create_sheet("Sheet1")

ws.cell(row=3, column=3, value=12345) # IPPIS (C3)
ws.cell(row=3, column=4, value="John Doe") # Name (D3)

wb.save("test_min_col2.xlsx")

wb_read = openpyxl.load_workbook("test_min_col2.xlsx", read_only=True, data_only=True)
ws_read = wb_read.active

for row in ws_read.iter_rows(min_row=3, max_row=ws_read.max_row, values_only=True):
    print("Values:", row)
    print("Length:", len(row))
