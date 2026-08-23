import openpyxl
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# Headers - skip column A
ws.cell(row=2, column=2, value="FILE NO")
ws.cell(row=2, column=3, value="IPPIS NUMBER")
ws.cell(row=2, column=4, value="NAME")

ws.cell(row=3, column=3, value=12345) # IPPIS
ws.cell(row=3, column=4, value="John Doe") # Name

wb.save("test_min_col.xlsx")

wb_read = openpyxl.load_workbook("test_min_col.xlsx", read_only=True, data_only=True)
ws_read = wb_read.active

for row in ws_read.iter_rows(min_row=3, max_row=ws_read.max_row):
    print("Cells:", [c.value for c in row])
    print("Length:", len(row))
