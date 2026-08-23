import openpyxl
from openpyxl import Workbook

# Create a test workbook
wb = Workbook()
ws = wb.active

# Headers
ws.append(["S/NO", "FILE NO", "IPPIS NUMBER", "NAME", "GL", "DEPT", "DIVISION"])

# Data with some missing cells
ws.cell(row=2, column=3, value=12345) # IPPIS
ws.cell(row=2, column=4, value="John Doe") # Name
ws.cell(row=2, column=5, value="GL08") # GL
# Missing other cells

wb.save("test.xlsx")

# Now read it
wb_read = openpyxl.load_workbook("test.xlsx", read_only=True, data_only=True)
ws_read = wb_read.active

for row in ws_read.iter_rows(min_row=2, max_row=ws_read.max_row):
    print("Cells:", [c.value for c in row])
    serial_no = row[0].value if len(row) > 0 else None
    ippis = row[2].value if len(row) > 2 else None
    name = row[3].value if len(row) > 3 else None
    print("serial_no:", serial_no, "ippis:", ippis, "name:", name)

# Test with values_only=True
print("With values_only=True:")
for row in ws_read.iter_rows(min_row=2, max_row=ws_read.max_row, values_only=True):
    print("Values:", row)
