import openpyxl
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.cell(row=2, column=2, value="FILE NO")
ws.cell(row=3, column=3, value=12345)
ws.cell(row=4, column=3, value=54321)
wb.save("test_none.xlsx")

wb_read = openpyxl.load_workbook("test_none.xlsx", read_only=True, data_only=True)
ws_read = wb_read.active

print("max_row:", ws_read.max_row)

rows = list(ws_read.iter_rows(min_row=3, max_row=None))
print("Iterated rows with max_row=None:", len(rows))
