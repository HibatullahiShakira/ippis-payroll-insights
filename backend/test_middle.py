import openpyxl
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
# Row 3: A has value, B is empty, C has value.
ws.cell(row=3, column=1, value="A_VAL")
ws.cell(row=3, column=3, value="C_VAL")
# No D, E, F, G

wb.save("test_middle.xlsx")

wb_read = openpyxl.load_workbook("test_middle.xlsx", read_only=True, data_only=True)
ws_read = wb_read.active

for row in ws_read.iter_rows(min_row=3, max_row=ws_read.max_row):
    print("row length:", len(row))
    print("row types:", [type(c) for c in row])
    try:
        print("values:", [c.value if hasattr(c, 'value') else 'NO_VALUE_ATTR' for c in row])
    except Exception as e:
        print("Error:", e)
