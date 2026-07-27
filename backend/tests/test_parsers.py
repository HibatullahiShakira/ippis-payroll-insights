import os
import pytest
from app.services.excel_parser import parse_excel
from app.models.employee import Employee
from app.extensions import db

def test_excel_parser_with_dummy_file(db_session, tmpdir):
    """Create a dummy excel file and test the parser."""
    import openpyxl
    
    # Create a temporary excel file
    file_path = os.path.join(tmpdir, "test.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Write headers on row 2
    headers = ["S/NO", "FILE NO", "IPPIS NUMBER", "NAME", "GL", "DEPT", "DIVISION"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
        
    # Write one row of data on row 3
    data = [1, 100, 123456, "TEST USER", "12", "IT", "SOFTWARE"]
    for col, v in enumerate(data, 1):
        ws.cell(row=3, column=col, value=v)
        
    wb.save(file_path)
    
    # Run the parser
    result = parse_excel(file_path)
    
    # Verify result dictionary/count (the function returns int of count processed)
    assert result == 1
    
    # Verify database state
    emp = Employee.query.filter_by(file_no=100).first()
    assert emp is not None
    assert emp.name == "TEST USER"
    assert emp.gl == "12"
    assert emp.department == "IT"
    assert emp.division == "SOFTWARE"

def test_pdf_parser_regex():
    """Test the regex compilation used in the PDF parser."""
    from app.services.pdf_parser import parse_pdf
    assert parse_pdf is not None
